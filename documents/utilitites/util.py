import requests
import json
import openpyxl
import time
from datetime import datetime

# -------------------------
# Настройки
# -------------------------
BITRIX_WEBHOOK_URL = "https://prav-buro.bitrix24.ru/rest/24/pa1x5irnfpbcnh27/"
REGION_FIELD_NAME = "UF_CRM_1745886887592"
CATEGORY_ID = 2
EXCEL_FILENAME = "deals_export.xlsx"


def bitrix(method: str, params: dict = None):
    """Универсальный запрос с паузой."""
    url = f"{BITRIX_WEBHOOK_URL}{method}"
    response = requests.get(url, params=params)
    time.sleep(0.25)
    return response.json()


# -------------------------------------------------------
# Получение телефона контакта
# -------------------------------------------------------
def fetch_contact_phone(contact_id: str | int) -> str:
    if not contact_id:
        return "—"

    print(f"[LOG] Загружаем контакт {contact_id}...")

    req = bitrix("crm.contact.get", {"ID": contact_id})

    if "result" not in req or not req["result"]:
        print(f"[LOG] Контакт {contact_id} не найден")
        return "—"

    phones = req["result"].get("PHONE", [])
    if not phones:
        return "—"

    phone = phones[0].get("VALUE", "").strip()
    print(f"[LOG] Телефон контакта {contact_id}: {phone}")

    return phone or "—"


def get_userfield_internal_id(field_name: str) -> str:
    print(f"[LOG] Запрашиваем список userfield...")
    data = bitrix("crm.deal.userfield.list")

    if "result" not in data:
        raise Exception(f"Bitrix error: {data}")

    for field in data["result"]:
        if field.get("FIELD_NAME") == field_name:
            print(f"[LOG] Нашли поле {field_name} → ID = {field['ID']}")
            return field["ID"]

    raise Exception(f"Поле {field_name} не найдено")


# -------------------------------------------------------
# Маппинг ENUM
# -------------------------------------------------------
def fetch_enum_map(field_name: str, save_to_file: str | None = None):
    internal_id = get_userfield_internal_id(field_name)

    print(f"[LOG] Запрашиваем ENUM ID = {internal_id}")
    data = bitrix("crm.deal.userfield.get", {"ID": internal_id})

    if "result" not in data:
        raise Exception(f"Bitrix error: {data}")

    enum_list = data["result"].get("LIST", [])
    mapping = {item["ID"]: item["VALUE"] for item in enum_list}

    print(f"[LOG] ENUM значений: {len(mapping)}")

    if save_to_file:
        with open(save_to_file, "w", encoding="utf-8") as f:
            json.dump(mapping, f, ensure_ascii=False, indent=4)

    return mapping


# -------------------------------------------------------
# Получение сделок
# -------------------------------------------------------
def fetch_all_deals(category_id: int):
    deals = []
    start = 0

    while True:
        print(f"[LOG] Загружаем сделки start={start}")

        params = {
            "start": start,
            "filter[CATEGORY_ID]": category_id,
            "select[]": [
                "ID",
                "TITLE",
                "CLOSEDATE",
                "CONTACT_ID",
                "UF_CRM_1745886887592",
                "UF_CRM_1754401876367",
                "UF_CRM_1745888327609",
                "UF_CRM_1746616466655",
            ],
        }

        data = bitrix("crm.deal.list", params)

        if "result" not in data:
            raise Exception(f"Ошибка Bitrix: {data}")

        batch = data["result"]
        print(f"[LOG] Получено {len(batch)} сделок")

        for d in batch:
            print(f"[LOG] Сделка ID={d.get('ID')} ФИО={d.get('TITLE')}")

        deals.extend(batch)

        if "next" in data:
            start = data["next"]
        else:
            break

    return deals


# -------------------------------------------------------
# Форматирование даты
# -------------------------------------------------------
def format_date(value):
    if not value:
        return ""

    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y")
    except:
        pass

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(value, fmt)
            return dt.strftime("%d.%m.%Y")
        except:
            pass

    try:
        ts = int(value)
        if ts < 10_000_000_000:
            dt = datetime.fromtimestamp(ts)
            return dt.strftime("%d.%m.%Y")
    except:
        pass

    return value


# -------------------------------------------------------
# Нормализация сделки
# -------------------------------------------------------
def normalize_deal(deal, region_map):
    deal_id = deal.get("ID", "")
    contact_id = deal.get("CONTACT_ID")

    phone = fetch_contact_phone(contact_id)

    return {
        "ID": deal_id,
        "ФИО": deal.get("TITLE", ""),
        "Телефон": phone,
        "Дата рождения": format_date(deal.get("UF_CRM_1745888327609", "")),
        "Фактический регион проживания": region_map.get(
            deal.get("UF_CRM_1745886887592", ""), ""
        ),
        "Сумма долга": deal.get("UF_CRM_1746616466655", ""),
        "Дата завершения": format_date(deal.get("CLOSEDATE", "")),
        "Дело на КадАрбитре": deal.get("UF_CRM_1754401876367", ""),
        "Ссылка на сделку": f"https://prav-buro.bitrix24.ru/crm/deal/details/{deal_id}/",
    }


# -------------------------------------------------------
# Экспорт в Excel
# -------------------------------------------------------
def export_to_excel(deals, region_map, filename="deals.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Сделки"

    headers = [
        "ID",
        "ФИО",
        "Телефон",
        "Дата рождения",
        "Фактический регион проживания",
        "Сумма долга",
        "Дата завершения",
        "Дело на КадАрбитре",
        "Ссылка на сделку",
    ]

    sheet.append(headers)

    for deal in deals:
        row = normalize_deal(deal, region_map)
        sheet.append([row[h] for h in headers])

    workbook.save(filename)
    print(f"[LOG] Файл сохранён: {filename}")



from google.oauth2 import service_account
from googleapiclient.discovery import build
import openpyxl

def upload_excel_to_google_sheets(
    excel_filename: str,
    spreadsheet_id: str,
    sheet_name: str,
    credentials_file: str = "creditnails-service.json"
):
    """Полная замена листа данными из Excel."""

    # --- 1. Авторизация ---
    creds = service_account.Credentials.from_service_account_file(
        credentials_file,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
    )

    service = build("sheets", "v4", credentials=creds)
    sheet = service.spreadsheets()

    # --- 2. Загружаем Excel ---
    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    # --- 3. Полная очистка листа ---
    sheet.values().clear(
        spreadsheetId=spreadsheet_id,
        range=f"{sheet_name}!A:Z"
    ).execute()

    # --- 4. Записываем данные ---
    sheet.values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{sheet_name}!A1",
        valueInputOption="RAW",
        body={"values": data}
    ).execute()

    print("Google Sheets успешно обновлён!")